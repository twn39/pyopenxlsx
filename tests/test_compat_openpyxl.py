import pytest
import datetime
import openpyxl
from pyopenxlsx import Workbook as PyWorkbook, Font, Fill, Border, Side, Alignment, XLPatternType, XLLineStyle, XLAlignmentStyle

def test_write_pyopenxlsx_read_openpyxl_datatypes(tmp_path):
    """
    Test Case 1a: Write data types with pyopenxlsx, Read with openpyxl.
    """
    filename = tmp_path / "compat_datatypes.xlsx"
    wb = PyWorkbook()
    ws = wb.active
    ws.title = "DataTypes"

    # 1. Strings
    ws["A1"].value = "Hello World"
    ws["A2"].value = "你好世界"  # Unicode
    ws["A3"].value = ""  # Empty string

    # 2. Integers
    ws["B1"].value = 42
    ws["B2"].value = -100
    ws["B3"].value = 0
    ws["B4"].value = 2147483647 # Max 32-bit signed

    # 3. Floats
    ws["C1"].value = 3.14159
    ws["C2"].value = -0.001
    ws["C3"].value = 1.23e10

    # 4. Booleans
    ws["D1"].value = True
    ws["D2"].value = False

    # 5. Dates / Datetimes
    # Note: openpyxl expects dates to be stored as numbers with a specific style, 
    # OR standard date format if no style. 
    # Just writing value usually results in Serial Date in Excel.
    # OpenPyXL reads these as numbers unless num_fmt is set.
    # pyopenxlsx might not auto-set num_fmt yet?
    dt = datetime.datetime(2023, 10, 27, 14, 30, 0)
    d = datetime.date(2023, 12, 25)
    
    ws["E1"].value = dt
    ws["E2"].value = d
    
    # Manually set style for dates so openpyxl recognizes them
    # (Assuming pyopenxlsx requires manual style app for now based on earlier exploration)
    date_style_idx = wb.add_style(number_format="yyyy-mm-dd hh:mm:ss")
    ws["E1"].style_index = date_style_idx
    
    date_only_style_idx = wb.add_style(number_format="yyyy-mm-dd")
    ws["E2"].style_index = date_only_style_idx

    wb.save(str(filename))

    # --- Verification with openpyxl ---
    wb_xl = openpyxl.load_workbook(filename)
    ws_xl = wb_xl["DataTypes"]

    # 1. Strings
    assert ws_xl["A1"].value == "Hello World"
    assert ws_xl["A2"].value == "你好世界"
    # openpyxl might read empty string as None if cell is empty, but if explicitly set to "", it might be "" or None depending on implementation.
    # Let's check what it actually is.
    # If pyopenxlsx wrote an empty string element, openpyxl usually sees implicit None or empty string.
    # We will adjust expectation if needed.
    assert ws_xl["A3"].value in ("", None) 

    # 2. Integers
    assert ws_xl["B1"].value == 42
    assert ws_xl["B2"].value == -100
    assert ws_xl["B3"].value == 0
    assert ws_xl["B4"].value == 2147483647

    # 3. Floats
    # Use approx for float comparison
    assert abs(ws_xl["C1"].value - 3.14159) < 1e-9
    assert abs(ws_xl["C2"].value - (-0.001)) < 1e-9
    assert abs(ws_xl["C3"].value - 1.23e10) < 1e-1

    # 4. Booleans
    assert ws_xl["D1"].value is True
    assert ws_xl["D2"].value is False

    # 5. Dates
    # Openpyxl should convert these to datetime objects if number format is correct
    assert isinstance(ws_xl["E1"].value, datetime.datetime)
    assert ws_xl["E1"].value == dt
    
    # Openpyxl often reads dates as datetimes with 00:00:00 time
    assert isinstance(ws_xl["E2"].value, datetime.datetime) # openpyxl reads dates as datetimes usually
    assert ws_xl["E2"].value.date() == d

    wb_xl.close()

def test_write_pyopenxlsx_read_openpyxl_styles(tmp_path):
    """
    Test Case 1b: Write styles with pyopenxlsx, Read with openpyxl.
    """
    filename = tmp_path / "compat_styles.xlsx"
    wb = PyWorkbook()
    ws = wb.active
    ws.title = "Styles"

    # 1. Font
    ws["A1"].value = "Bold"
    ws["B1"].value = "Italic" 
    ws["C1"].value = "Red"
    ws["D1"].value = "Size 14"

    # Font args: name, size, bold, italic, underline, strikethrough, color
    # Colors in pyopenxlsx are usually ARGB hex strings
    font_bold = Font(bold=True)
    font_italic = Font(italic=True)
    font_red = Font(color="FF0000") 
    font_size = Font(size=14)

    style_bold = wb.add_style(font=font_bold)
    style_italic = wb.add_style(font=font_italic)
    style_red = wb.add_style(font=font_red)
    style_size = wb.add_style(font=font_size)

    ws["A1"].style_index = style_bold
    ws["B1"].style_index = style_italic
    ws["C1"].style_index = style_red
    ws["D1"].style_index = style_size

    # 2. Fill
    ws["A2"].value = "Solid Fill"
    # Fill args: pattern_type, color (fg), background_color (bg)
    fill_solid = Fill(pattern_type="solid", color="FFFF00") # Yellow
    style_fill = wb.add_style(fill=fill_solid)
    ws["A2"].style_index = style_fill

    # 3. Border
    ws["A3"].value = "Bordered"
    # Border args: left, right, top, bottom, diagonal...
    # Side args: style, color
    thin_side = Side(style="thin", color="000000")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    style_border = wb.add_style(border=border)
    ws["A3"].style_index = style_border

    # 4. Alignment
    ws["A4"].value = "Centered"
    # Alignment args: horizontal, vertical, wrap_text, rotation...
    align = Alignment(horizontal="center", vertical="center")
    style_align = wb.add_style(alignment=align)
    ws["A4"].style_index = style_align

    wb.save(str(filename))

    # --- Verification ---
    wb_xl = openpyxl.load_workbook(filename)
    ws_xl = wb_xl["Styles"]

    # Font
    assert ws_xl["A1"].font.b is True
    assert ws_xl["B1"].font.i is True
    # Check color. openpyxl might return "00FF0000" (ARGB) or valid Theme color.
    # pyopenxlsx "FF0000" -> ARGB "FFFF0000".
    if ws_xl["C1"].font.color and ws_xl["C1"].font.color.rgb:
         # Accept either FFFF0000 or just FF0000 logic depending on normalization
         assert ws_xl["C1"].font.color.rgb.upper() == "FFFF0000"
    
    assert ws_xl["D1"].font.sz == 14

    # Fill
    assert ws_xl["A2"].fill.patternType == "solid"
    if ws_xl["A2"].fill.fgColor:
        assert ws_xl["A2"].fill.fgColor.rgb.upper() == "FFFFFF00"

    # Border
    assert ws_xl["A3"].border.left.style == "thin"
    assert ws_xl["A3"].border.right.style == "thin"
    assert ws_xl["A3"].border.top.style == "thin"
    assert ws_xl["A3"].border.bottom.style == "thin"

    # Alignment
    assert ws_xl["A4"].alignment.horizontal == "center"
    assert ws_xl["A4"].alignment.vertical == "center"

    wb_xl.close()


def test_write_pyopenxlsx_read_openpyxl_images(tmp_path):
    """
    Test Case 1c: Write images with pyopenxlsx, Read with openpyxl.
    """
    filename = tmp_path / "compat_images.xlsx"
    img_path = tmp_path / "test_image.png"
    
    # Create a red 1x1 PNG
    img_data = b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\x0cIDAT\x08\xd7c\xf8\xff\xff?0\x00\x03\xff\x01\xfe\x8e\xfe\x1d\x00\x00\x00\x00IEND\xaeB`\x82"
    with open(img_path, "wb") as f:
        f.write(img_data)

    wb = PyWorkbook()
    ws = wb.active
    ws.title = "Images"
    
    # Add image
    ws.add_image(str(img_path), anchor="C3", width=100, height=100)
    
    wb.save(str(filename))

    # --- Verification ---
    wb_xl = openpyxl.load_workbook(filename)
    ws_xl = wb_xl["Images"]
    
    # openpyxl stores images in ws._images
    assert len(ws_xl._images) == 1
    
    wb_xl.close()


def test_write_openpyxl_read_pyopenxlsx_datatypes_and_images(tmp_path):
    """
    Test Case 2: Write with openpyxl, Read with pyopenxlsx.
    """
    filename = tmp_path / "reverse_compat.xlsx"
    
    wb_xl = openpyxl.Workbook()
    ws_xl = wb_xl.active
    ws_xl.title = "Reverse"
    
    ws_xl["A1"] = "Created by OpenPyXL"
    ws_xl["B1"] = 999
    ws_xl["C1"] = 123.456
    
    # Add an image
    img_path = tmp_path / "xl_image.png"
    img_data = b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\x0cIDAT\x08\xd7c\xf8\xff\xff?0\x00\x03\xff\x01\xfe\x8e\xfe\x1d\x00\x00\x00\x00IEND\xaeB`\x82"
    with open(img_path, "wb") as f:
        f.write(img_data)
        
    xl_img = openpyxl.drawing.image.Image(img_path)
    # Anchor to D4
    ws_xl.add_image(xl_img, "D4")
    
    wb_xl.save(filename)
    
    # --- Verification ---
    wb = PyWorkbook(str(filename))
    ws = wb.active
    
    assert ws.title == "Reverse"
    assert ws["A1"].value == "Created by OpenPyXL"
    assert ws["B1"].value == 999
    assert abs(ws["C1"].value - 123.456) < 1e-9
    
    # Check images
    images = wb.get_embedded_images()
    # Depending on how openpyxl writes images, they should be typically extractable
    assert len(images) >= 1
    
    wb.close()

def test_write_pyopenxlsx_read_openpyxl_formulas(tmp_path):
    """
    Test Case 1d: Write formulas with pyopenxlsx, Read with openpyxl.
    """
    filename = tmp_path / "compat_formulas.xlsx"
    wb = PyWorkbook()
    ws = wb.active
    ws.title = "Formulas"

    # Data for formulas
    ws["A1"].value = 10
    ws["A2"].value = 20
    ws["A3"].value = 30

    ws["B1"].formula = "SUM(A1:A3)"
    ws["B2"].formula = "AVERAGE(A1:A3)"
    ws["B3"].formula = "A1+A2"
    
    wb.save(str(filename))

    # --- Verification ---
    wb_xl = openpyxl.load_workbook(filename, data_only=False)
    ws_xl = wb_xl["Formulas"]

    # openpyxl reads data_only=False by default, so .value should be the formula string
    assert ws_xl["B1"].value == "=SUM(A1:A3)"
    assert ws_xl["B2"].value == "=AVERAGE(A1:A3)"
    # Some libs might capitalize or change whitespace, but simple ones usually match.
    # openpyxl adds '=' if it's missing in the file but usually it is stored with '=' in XML.
    assert ws_xl["B3"].value == "=A1+A2"

    wb_xl.close()


def test_write_openpyxl_read_pyopenxlsx_formulas(tmp_path):
    """
    Test Case 2b: Write formulas with openpyxl, Read with pyopenxlsx.
    """
    filename = tmp_path / "reverse_compat_formulas.xlsx"
    
    wb_xl = openpyxl.Workbook()
    ws_xl = wb_xl.active
    ws_xl.title = "ReverseFormulas"
    
    ws_xl["A1"] = 5
    ws_xl["A2"] = 15
    
    ws_xl["B1"] = "=SUM(A1:A2)"
    ws_xl["B2"] = "=MAX(A1:A2)"
    
    wb_xl.save(filename)
    
    # --- Verification ---
    wb = PyWorkbook(str(filename))
    ws = wb.active
    
    # Check formulas
    # Assuming .formula property exists and returns string
    # openpyxl stores formula with '=', pyopenxlsx might return with or without depending on impl.
    # Standardize on expecting '=' or checking containment.
    
    f1 = ws["B1"].formula
    f2 = ws["B2"].formula
    
    # It might return "SUM(A1:A2)" or "=SUM(A1:A2)"
    # Let's handle both for robustness or assert exact if we know. 
    # OpenXLSX usually returns the formula string.
    assert "SUM(A1:A2)" in str(f1).upper()
    assert "MAX(A1:A2)" in str(f2).upper()
    
    wb.close()

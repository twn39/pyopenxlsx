"""Performance benchmarks for pyopenxlsx vs openpyxl and internal paths.

Markers
-------
- ``benchmark``: all timed cases (pytest-benchmark)
- ``benchmark_fast``: default developer suite (~large scale, no 1M extreme)
- ``benchmark_extreme``: ~1M-cell write cases (slow; opt-in)

Examples::

    # Fast suite only
    uv run pytest tests/test_benchmark.py -m benchmark_fast --benchmark-only -q

    # Extreme write paths
    uv run pytest tests/test_benchmark.py -m benchmark_extreme --benchmark-only -q

    # Export JSON for comparison
    uv run pytest tests/test_benchmark.py -m benchmark_fast \\
        --benchmark-only --benchmark-json=benchmark.json
    uv run python scripts/compare_benchmarks.py baseline.json benchmark.json
"""

from __future__ import annotations

import asyncio
import os
import time
from datetime import date, datetime, timedelta
from typing import Any, Callable, Iterable, List, Sequence

import numpy as np
import openpyxl
import pytest
from pyopenxlsx import Workbook as PyWorkbook
from pyopenxlsx import load_workbook_async

psutil = pytest.importorskip("psutil")

# ---------------------------------------------------------------------------
# Markers helpers
# ---------------------------------------------------------------------------

fast = pytest.mark.benchmark_fast
extreme = pytest.mark.benchmark_extreme


# ---------------------------------------------------------------------------
# Resource monitoring (measures the callable itself)
# ---------------------------------------------------------------------------


class ResourceMonitor:
    """Context manager that reports wall time, RSS delta, and CPU for a block."""

    def __init__(self, name: str):
        self.name = name
        self.process = psutil.Process(os.getpid())
        self.peak_rss_mb = 0.0

    def __enter__(self):
        self.start_mem = self.process.memory_info().rss / (1024 * 1024)
        self.start_cpu = self.process.cpu_times()
        self.start_time = time.perf_counter()
        self.peak_rss_mb = self.start_mem
        return self

    def sample(self) -> None:
        rss = self.process.memory_info().rss / (1024 * 1024)
        if rss > self.peak_rss_mb:
            self.peak_rss_mb = rss

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.sample()
        end_time = time.perf_counter()
        end_mem = self.process.memory_info().rss / (1024 * 1024)
        end_cpu = self.process.cpu_times()
        duration = end_time - self.start_time
        mem_diff = end_mem - self.start_mem
        cpu_user = end_cpu.user - self.start_cpu.user
        cpu_sys = end_cpu.system - self.start_cpu.system
        cpu_total = cpu_user + cpu_sys
        print(f"\n[Resource: {self.name}]")
        print(f"  Time: {duration:.4f}s")
        print(f"  Memory Delta: {mem_diff:+.2f} MB (End: {end_mem:.2f} MB)")
        print(f"  Peak RSS (approx): {self.peak_rss_mb:.2f} MB")
        print(
            f"  CPU Usage: User {cpu_user:.4f}s, System {cpu_sys:.4f}s, "
            f"Total {cpu_total:.4f}s"
        )
        if duration > 0:
            print(f"  CPU Load: {(cpu_total / duration) * 100:.1f}%")


def _monitored(name: str, fn: Callable[..., Any], *args, **kwargs) -> Any:
    """Run *fn* under ResourceMonitor (for use inside benchmark())."""
    with ResourceMonitor(name):
        return fn(*args, **kwargs)


# ---------------------------------------------------------------------------
# Fixtures — data
# ---------------------------------------------------------------------------


@pytest.fixture
def temp_xlsx_file(tmp_path):
    return str(tmp_path / "benchmark.xlsx")


@pytest.fixture(scope="session")
def small_str_data() -> List[List[str]]:
    """100×10 strings — fair small write."""
    return [[f"R{r}C{c}" for c in range(1, 11)] for r in range(1, 101)]


@pytest.fixture(scope="session")
def large_str_data() -> List[List[str]]:
    """1000×50 strings."""
    return [[f"R{r}C{c}" for c in range(1, 51)] for r in range(1, 1001)]


@pytest.fixture(scope="session")
def large_float_list() -> List[List[float]]:
    """1000×50 floats as Python lists (same logical grid as numpy)."""
    return [
        [float(r * 50 + c) for c in range(50)] for r in range(1000)
    ]


@pytest.fixture(scope="session")
def large_np_data():
    """1000×50 float64 ndarray."""
    return np.arange(1000 * 50, dtype=np.float64).reshape(1000, 50)


@pytest.fixture(scope="session")
def extreme_str_data() -> List[List[str]]:
    """10000×100 strings (~1M cells)."""
    return [[f"R{r}C{c}" for c in range(1, 101)] for r in range(1, 10001)]


@pytest.fixture(scope="session")
def extreme_float_list() -> List[List[float]]:
    return [
        [float(r * 100 + c) for c in range(100)] for r in range(10000)
    ]


@pytest.fixture(scope="session")
def extreme_np_data():
    return np.arange(10000 * 100, dtype=np.float64).reshape(10000, 100)


@pytest.fixture(scope="session")
def large_date_data() -> List[List[Any]]:
    """500×20 mix of date/datetime for coercion cost measurement."""
    base = date(2020, 1, 1)
    rows: List[List[Any]] = []
    for r in range(500):
        row: List[Any] = []
        for c in range(20):
            d = base + timedelta(days=r, hours=c % 24)
            row.append(d if c % 2 == 0 else datetime.combine(d, datetime.min.time()) + timedelta(hours=c))
        rows.append(row)
    return rows


@pytest.fixture(scope="session")
def large_mixed_data() -> List[List[Any]]:
    """1000×20 mixed types (str / int / float / None / date)."""
    base = date(2023, 6, 1)
    rows: List[List[Any]] = []
    for r in range(1000):
        rows.append(
            [
                f"id-{r}",
                r,
                float(r) * 1.5,
                None if r % 7 == 0 else r % 2 == 0,
                base + timedelta(days=r % 365),
                f"note-{r % 10}",
                r * 0.01,
                None,
                "x",
                r + 1,
            ]
            * 2  # 20 cols
        )
    return rows


# ---------------------------------------------------------------------------
# Fixtures — files
# ---------------------------------------------------------------------------


@pytest.fixture
def large_file(tmp_path):
    """Shared 1000×20 string workbook for fair read/iterate (openpyxl-generated)."""
    filepath = str(tmp_path / "large_input.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in range(1, 1001):
        for c in range(1, 21):
            ws.cell(row=r, column=c, value=f"Val_{r}_{c}")
    wb.save(filepath)
    return filepath


@pytest.fixture
def multiple_files(tmp_path):
    files = []
    data = np.arange(100 * 50, dtype=np.float64).reshape(100, 50)
    for i in range(10):
        fp = str(tmp_path / f"async_bench_{i}.xlsx")
        wb = PyWorkbook()
        try:
            ws = wb.active
            ws.write_range(1, 1, data)
            wb.save(fp)
        finally:
            wb.close()
        files.append(fp)
    return files


@pytest.fixture
def output_dir(tmp_path):
    d = tmp_path / "async_out"
    d.mkdir()
    return d


# ---------------------------------------------------------------------------
# Write helpers (always close)
# ---------------------------------------------------------------------------


def write_py_cell(data: Sequence[Sequence[Any]], filepath: str) -> None:
    wb = PyWorkbook()
    try:
        ws = wb.active
        for r, row in enumerate(data, 1):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c).value = val
        wb.save(filepath)
    finally:
        wb.close()


def write_ox_cell(data: Sequence[Sequence[Any]], filepath: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    for r, row in enumerate(data, 1):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    wb.save(filepath)


def write_py_set_cell(data: Sequence[Sequence[Any]], filepath: str) -> None:
    wb = PyWorkbook()
    try:
        ws = wb.active
        for r, row in enumerate(data, 1):
            for c, val in enumerate(row, 1):
                ws.set_cell_value(r, c, val)
        wb.save(filepath)
    finally:
        wb.close()


def write_py_rows(data: Sequence[Sequence[Any]], filepath: str) -> None:
    wb = PyWorkbook()
    try:
        ws = wb.active
        ws.write_rows(1, data)
        wb.save(filepath)
    finally:
        wb.close()


def write_py_set_cells(data: Sequence[Sequence[Any]], filepath: str) -> None:
    wb = PyWorkbook()
    try:
        ws = wb.active
        cells = [
            (r, c, val)
            for r, row in enumerate(data, 1)
            for c, val in enumerate(row, 1)
        ]
        ws.set_cells(cells)
        wb.save(filepath)
    finally:
        wb.close()


def write_py_numpy(np_data, filepath: str) -> None:
    wb = PyWorkbook()
    try:
        ws = wb.active
        ws.write_range(1, 1, np_data)
        wb.save(filepath)
    finally:
        wb.close()


def write_py_stream(
    data: Sequence[Sequence[Any]], filepath: str, *, auto_dates: bool = True
) -> None:
    wb = PyWorkbook()
    try:
        wb.auto_date_formats = auto_dates
        ws = wb.active
        with ws.stream_writer() as writer:
            for row in data:
                writer.append_row(row)
        wb.save(filepath)
    finally:
        wb.close()


def write_py_rows_dates(
    data: Sequence[Sequence[Any]], filepath: str, *, auto_dates: bool = True
) -> None:
    wb = PyWorkbook()
    try:
        wb.auto_date_formats = auto_dates
        ws = wb.active
        ws.write_rows(1, data)
        wb.save(filepath)
    finally:
        wb.close()


def write_py_cell_dates(
    data: Sequence[Sequence[Any]], filepath: str, *, auto_dates: bool = True
) -> None:
    wb = PyWorkbook()
    try:
        wb.auto_date_formats = auto_dates
        ws = wb.active
        for r, row in enumerate(data, 1):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c).value = val
        wb.save(filepath)
    finally:
        wb.close()


def write_py_dataframe(df, filepath: str) -> None:
    wb = PyWorkbook()
    try:
        ws = wb.active
        ws.write_dataframe(df, header=True, index=False)
        wb.save(filepath)
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Read / iterate helpers
# ---------------------------------------------------------------------------


def read_point_py(filepath: str) -> Any:
    wb = PyWorkbook(filepath)
    try:
        return wb.active.cell(row=500, column=10).value
    finally:
        wb.close()


def read_point_ox(filepath: str) -> Any:
    wb = openpyxl.load_workbook(filepath, data_only=False)
    try:
        return wb.active.cell(row=500, column=10).value
    finally:
        wb.close()


def read_scan_stream_py(filepath: str) -> Any:
    wb = PyWorkbook(filepath)
    try:
        val = None
        with wb.active.stream_reader() as reader:
            for i, row in enumerate(reader, 1):
                if i == 500:
                    val = row[9]
                    break
        return val
    finally:
        wb.close()


def read_scan_bulk_py(filepath: str) -> int:
    wb = PyWorkbook(filepath)
    try:
        data = wb.active.get_rows_data()
        return sum(len(row) for row in data)
    finally:
        wb.close()


def iterate_cell_py(filepath: str) -> int:
    wb = PyWorkbook(filepath)
    try:
        ws = wb.active
        count = 0
        for r in range(1, 1001):
            for c in range(1, 21):
                _ = ws.cell(row=r, column=c).value
                count += 1
        return count
    finally:
        wb.close()


def iterate_cell_ox(filepath: str) -> int:
    wb = openpyxl.load_workbook(filepath)
    try:
        count = 0
        for row in wb.active.iter_rows(min_row=1, max_row=1000, min_col=1, max_col=20):
            for cell in row:
                _ = cell.value
                count += 1
        return count
    finally:
        wb.close()


def iterate_values_py(filepath: str) -> int:
    wb = PyWorkbook(filepath)
    try:
        count = 0
        for row in wb.active.iter_rows(
            min_row=1, max_row=1000, min_col=1, max_col=20, values_only=True
        ):
            count += len(row)
        return count
    finally:
        wb.close()


def iterate_values_ox(filepath: str) -> int:
    wb = openpyxl.load_workbook(filepath, read_only=True)
    try:
        count = 0
        for row in wb.active.iter_rows(
            min_row=1, max_row=1000, min_col=1, max_col=20, values_only=True
        ):
            count += len(row)
        return count
    finally:
        wb.close()


def iterate_stream_py(filepath: str) -> int:
    wb = PyWorkbook(filepath)
    try:
        count = 0
        with wb.active.stream_reader() as reader:
            for row in reader:
                count += len(row)
        return count
    finally:
        wb.close()


def iterate_bulk_py(filepath: str) -> int:
    return read_scan_bulk_py(filepath)


# ---------------------------------------------------------------------------
# WRITE — small strings (fair)
# ---------------------------------------------------------------------------


@fast
@pytest.mark.benchmark(group="write_small_str")
def test_write_small_str_py_cell(benchmark, temp_xlsx_file, small_str_data):
    benchmark(write_py_cell, small_str_data, temp_xlsx_file)


@fast
@pytest.mark.benchmark(group="write_small_str")
def test_write_small_str_ox_cell(benchmark, temp_xlsx_file, small_str_data):
    benchmark(write_ox_cell, small_str_data, temp_xlsx_file)


@fast
@pytest.mark.benchmark(group="write_small_str")
def test_write_small_str_py_rows(benchmark, temp_xlsx_file, small_str_data):
    benchmark(write_py_rows, small_str_data, temp_xlsx_file)


# ---------------------------------------------------------------------------
# WRITE — large strings (fair across libraries / paths)
# ---------------------------------------------------------------------------


@fast
@pytest.mark.benchmark(group="write_large_str")
def test_write_large_str_py_cell(benchmark, temp_xlsx_file, large_str_data):
    benchmark(write_py_cell, large_str_data, temp_xlsx_file)


@fast
@pytest.mark.benchmark(group="write_large_str")
def test_write_large_str_ox_cell(benchmark, temp_xlsx_file, large_str_data):
    benchmark(write_ox_cell, large_str_data, temp_xlsx_file)


@fast
@pytest.mark.benchmark(group="write_large_str")
def test_write_large_str_py_set_cell(benchmark, temp_xlsx_file, large_str_data):
    benchmark(write_py_set_cell, large_str_data, temp_xlsx_file)


@fast
@pytest.mark.benchmark(group="write_large_str")
def test_write_large_str_py_rows(benchmark, temp_xlsx_file, large_str_data):
    benchmark(write_py_rows, large_str_data, temp_xlsx_file)


@fast
@pytest.mark.benchmark(group="write_large_str")
def test_write_large_str_py_set_cells(benchmark, temp_xlsx_file, large_str_data):
    benchmark(write_py_set_cells, large_str_data, temp_xlsx_file)


@fast
@pytest.mark.benchmark(group="write_large_str")
def test_write_large_str_py_stream(benchmark, temp_xlsx_file, large_str_data):
    benchmark(write_py_stream, large_str_data, temp_xlsx_file)


# ---------------------------------------------------------------------------
# WRITE — large floats (fair: list vs numpy)
# ---------------------------------------------------------------------------


@fast
@pytest.mark.benchmark(group="write_large_float")
def test_write_large_float_py_rows(benchmark, temp_xlsx_file, large_float_list):
    benchmark(write_py_rows, large_float_list, temp_xlsx_file)


@fast
@pytest.mark.benchmark(group="write_large_float")
def test_write_large_float_py_numpy(benchmark, temp_xlsx_file, large_np_data):
    benchmark(write_py_numpy, large_np_data, temp_xlsx_file)


@fast
@pytest.mark.benchmark(group="write_large_float")
def test_write_large_float_ox_cell(benchmark, temp_xlsx_file, large_float_list):
    benchmark(write_ox_cell, large_float_list, temp_xlsx_file)


@fast
@pytest.mark.benchmark(group="write_large_float")
def test_write_large_float_py_stream(benchmark, temp_xlsx_file, large_float_list):
    benchmark(write_py_stream, large_float_list, temp_xlsx_file)


# ---------------------------------------------------------------------------
# WRITE — dates (coercion + auto_date_formats)
# ---------------------------------------------------------------------------


@fast
@pytest.mark.benchmark(group="write_large_dates")
def test_write_dates_py_cell_auto_on(benchmark, temp_xlsx_file, large_date_data):
    benchmark(write_py_cell_dates, large_date_data, temp_xlsx_file, auto_dates=True)


@fast
@pytest.mark.benchmark(group="write_large_dates")
def test_write_dates_py_cell_auto_off(benchmark, temp_xlsx_file, large_date_data):
    benchmark(write_py_cell_dates, large_date_data, temp_xlsx_file, auto_dates=False)


@fast
@pytest.mark.benchmark(group="write_large_dates")
def test_write_dates_py_rows_auto_on(benchmark, temp_xlsx_file, large_date_data):
    benchmark(write_py_rows_dates, large_date_data, temp_xlsx_file, auto_dates=True)


@fast
@pytest.mark.benchmark(group="write_large_dates")
def test_write_dates_py_rows_auto_off(benchmark, temp_xlsx_file, large_date_data):
    benchmark(write_py_rows_dates, large_date_data, temp_xlsx_file, auto_dates=False)


@fast
@pytest.mark.benchmark(group="write_large_dates")
def test_write_dates_py_stream_auto_on(benchmark, temp_xlsx_file, large_date_data):
    benchmark(write_py_stream, large_date_data, temp_xlsx_file, auto_dates=True)


@fast
@pytest.mark.benchmark(group="write_large_dates")
def test_write_dates_py_stream_auto_off(benchmark, temp_xlsx_file, large_date_data):
    benchmark(write_py_stream, large_date_data, temp_xlsx_file, auto_dates=False)


# ---------------------------------------------------------------------------
# WRITE — mixed types
# ---------------------------------------------------------------------------


@fast
@pytest.mark.benchmark(group="write_large_mixed")
def test_write_mixed_py_rows(benchmark, temp_xlsx_file, large_mixed_data):
    benchmark(write_py_rows, large_mixed_data, temp_xlsx_file)


@fast
@pytest.mark.benchmark(group="write_large_mixed")
def test_write_mixed_py_stream(benchmark, temp_xlsx_file, large_mixed_data):
    benchmark(write_py_stream, large_mixed_data, temp_xlsx_file)


@fast
@pytest.mark.benchmark(group="write_large_mixed")
def test_write_mixed_ox_cell(benchmark, temp_xlsx_file, large_mixed_data):
    benchmark(write_ox_cell, large_mixed_data, temp_xlsx_file)


# ---------------------------------------------------------------------------
# WRITE — extreme (~1M cells), opt-in
# ---------------------------------------------------------------------------


@extreme
@pytest.mark.benchmark(group="write_extreme_float")
def test_extreme_float_py_numpy(benchmark, temp_xlsx_file, extreme_np_data):
    benchmark(
        _monitored, "py_numpy_1M", write_py_numpy, extreme_np_data, temp_xlsx_file
    )


@extreme
@pytest.mark.benchmark(group="write_extreme_float")
def test_extreme_float_py_rows(benchmark, temp_xlsx_file, extreme_float_list):
    benchmark(
        _monitored, "py_rows_float_1M", write_py_rows, extreme_float_list, temp_xlsx_file
    )


@extreme
@pytest.mark.benchmark(group="write_extreme_str")
def test_extreme_str_py_rows(benchmark, temp_xlsx_file, extreme_str_data):
    benchmark(
        _monitored, "py_rows_str_1M", write_py_rows, extreme_str_data, temp_xlsx_file
    )


@extreme
@pytest.mark.benchmark(group="write_extreme_str")
def test_extreme_str_py_stream(benchmark, temp_xlsx_file, extreme_str_data):
    benchmark(
        _monitored, "py_stream_str_1M", write_py_stream, extreme_str_data, temp_xlsx_file
    )


@extreme
@pytest.mark.benchmark(group="write_extreme_str")
def test_extreme_str_ox_cell(benchmark, temp_xlsx_file, extreme_str_data):
    benchmark(
        _monitored, "ox_cell_str_1M", write_ox_cell, extreme_str_data, temp_xlsx_file
    )


# ---------------------------------------------------------------------------
# READ — point vs scan (separate groups)
# ---------------------------------------------------------------------------


@fast
@pytest.mark.benchmark(group="read_point")
def test_read_point_py(benchmark, large_file):
    benchmark(read_point_py, large_file)


@fast
@pytest.mark.benchmark(group="read_point")
def test_read_point_ox(benchmark, large_file):
    benchmark(read_point_ox, large_file)


@fast
@pytest.mark.benchmark(group="read_scan")
def test_read_scan_stream_py(benchmark, large_file):
    benchmark(read_scan_stream_py, large_file)


@fast
@pytest.mark.benchmark(group="read_scan")
def test_read_scan_bulk_py(benchmark, large_file):
    benchmark(read_scan_bulk_py, large_file)


# ---------------------------------------------------------------------------
# ITERATE — cell path vs values_only (symmetric)
# ---------------------------------------------------------------------------


@fast
@pytest.mark.benchmark(group="iterate_cell")
def test_iterate_cell_py(benchmark, large_file):
    benchmark(iterate_cell_py, large_file)


@fast
@pytest.mark.benchmark(group="iterate_cell")
def test_iterate_cell_ox(benchmark, large_file):
    benchmark(iterate_cell_ox, large_file)


@fast
@pytest.mark.benchmark(group="iterate_values")
def test_iterate_values_py(benchmark, large_file):
    benchmark(iterate_values_py, large_file)


@fast
@pytest.mark.benchmark(group="iterate_values")
def test_iterate_values_ox(benchmark, large_file):
    benchmark(iterate_values_ox, large_file)


@fast
@pytest.mark.benchmark(group="iterate_values")
def test_iterate_bulk_py(benchmark, large_file):
    benchmark(iterate_bulk_py, large_file)


@fast
@pytest.mark.benchmark(group="iterate_values")
def test_iterate_stream_py(benchmark, large_file):
    benchmark(iterate_stream_py, large_file)


# ---------------------------------------------------------------------------
# LOAD
# ---------------------------------------------------------------------------


@fast
@pytest.mark.benchmark(group="load")
def test_load_py(benchmark, large_file):
    def load():
        wb = PyWorkbook(large_file)
        wb.close()

    benchmark(load)


@fast
@pytest.mark.benchmark(group="load")
def test_load_ox(benchmark, large_file):
    def load():
        wb = openpyxl.load_workbook(large_file)
        wb.close()

    benchmark(load)


@fast
@pytest.mark.benchmark(group="load")
def test_load_ox_readonly(benchmark, large_file):
    def load():
        wb = openpyxl.load_workbook(large_file, read_only=True)
        wb.close()

    benchmark(load)


# ---------------------------------------------------------------------------
# PANDAS (optional)
# ---------------------------------------------------------------------------


@fast
@pytest.mark.benchmark(group="write_dataframe")
def test_write_dataframe_py(benchmark, temp_xlsx_file, large_np_data):
    pd = pytest.importorskip("pandas")
    df = pd.DataFrame(large_np_data)
    benchmark(write_py_dataframe, df, temp_xlsx_file)


# ---------------------------------------------------------------------------
# ASYNC / concurrent (multi-file; educational, not single-book thread safety)
# ---------------------------------------------------------------------------


def read_files_sync(files: Iterable[str]) -> None:
    for fp in files:
        wb = PyWorkbook(fp)
        try:
            _ = wb.active.get_range_data(1, 1, 100, 50)
        finally:
            wb.close()


async def read_files_async(files: Sequence[str]) -> None:
    async def read_one(fp: str) -> None:
        wb = await load_workbook_async(fp)
        try:
            await wb.active.get_range_data_async(1, 1, 100, 50)
        finally:
            await wb.close_async()

    await asyncio.gather(*(read_one(fp) for fp in files))


def write_files_sync(output_dir) -> None:
    data = np.arange(100 * 50, dtype=np.float64).reshape(100, 50)
    for i in range(10):
        fp = str(output_dir / f"sync_out_{i}.xlsx")
        write_py_numpy(data, fp)


async def write_files_async(output_dir) -> None:
    data = np.arange(100 * 50, dtype=np.float64).reshape(100, 50)

    async def write_one(i: int) -> None:
        fp = str(output_dir / f"async_out_{i}.xlsx")
        wb = PyWorkbook()
        try:
            await wb.active.write_range_async(1, 1, data)
            await wb.save_async(fp)
        finally:
            await wb.close_async()

    await asyncio.gather(*(write_one(i) for i in range(10)))


def run_async(benchmark, func, *args) -> None:
    def wrapper():
        asyncio.run(func(*args))

    benchmark(wrapper)


@fast
@pytest.mark.benchmark(group="async_read_multifile")
def test_async_read_sync_loop(benchmark, multiple_files):
    """Baseline: sequential multi-file read."""
    benchmark(read_files_sync, multiple_files)


@fast
@pytest.mark.benchmark(group="async_read_multifile")
def test_async_read_gather(benchmark, multiple_files):
    """Concurrent multi-file read via asyncio.gather + to_thread helpers."""
    run_async(benchmark, read_files_async, multiple_files)


@fast
@pytest.mark.benchmark(group="async_write_multifile")
def test_async_write_sync_loop(benchmark, output_dir):
    benchmark(write_files_sync, output_dir)


@fast
@pytest.mark.benchmark(group="async_write_multifile")
def test_async_write_gather(benchmark, output_dir):
    run_async(benchmark, write_files_async, output_dir)


@fast
@pytest.mark.benchmark(group="async_loop_write_multifile")
def test_async_loop_write_sync(benchmark, output_dir, small_str_data):
    def run(out, data):
        for i in range(5):
            write_py_cell(data, str(out / f"loop_sync_{i}.xlsx"))

    benchmark(run, output_dir, small_str_data)


@fast
@pytest.mark.benchmark(group="async_loop_write_multifile")
def test_async_loop_write_gather(benchmark, output_dir, small_str_data):
    async def run(out, data):
        async def one(i):
            await asyncio.to_thread(
                write_py_cell, data, str(out / f"loop_async_{i}.xlsx")
            )

        await asyncio.gather(*(one(i) for i in range(5)))

    run_async(benchmark, run, output_dir, small_str_data)
